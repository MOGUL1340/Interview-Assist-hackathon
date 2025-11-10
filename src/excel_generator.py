import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging
import os
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def create_interview_excel(interview_plan, code_challenges, output_path=None):
    """
    Create a comprehensive Excel file with interview plan and evaluation form.

    Args:
        interview_plan (dict): Complete interview plan
        code_challenges (dict): Code challenges suite
        output_path (str): Output file path (optional)

    Returns:
        str: Path to the generated Excel file
    """
    logging.info("Creating Excel interview plan")

    try:
        # Create workbook
        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Create sheets
        create_overview_sheet(wb, interview_plan)
        topic_score_ranges = create_questions_sheet(wb, interview_plan)
        create_evaluation_sheet(wb, interview_plan, topic_score_ranges)
        create_code_challenges_sheet(wb, code_challenges)
        create_notes_sheet(wb, interview_plan)

        # Generate output path if not provided
        if not output_path:
            candidate_name = interview_plan.get("metadata", {}).get("candidate_name", "Candidate")
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = f"interview_plan_{candidate_name.replace(' ', '_')}_{timestamp}.xlsx"

        # Ensure output directory exists
        output_dir = os.path.dirname(output_path) if os.path.dirname(output_path) else "."
        if not os.path.exists(output_dir) and output_dir != ".":
            os.makedirs(output_dir)

        # Save workbook
        wb.save(output_path)
        logging.info(f"Excel file created successfully: {output_path}")

        return output_path

    except Exception as e:
        logging.error(f"Error creating Excel file: {str(e)}")
        return None

def create_overview_sheet(wb, interview_plan):
    """Create overview sheet with interview summary."""
    ws = wb.create_sheet("Overview", 0)

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "INTERVIEW PLAN - OVERVIEW"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Metadata
    metadata = interview_plan.get("metadata", {})
    row = 3

    metadata_items = [
        ("Candidate:", metadata.get("candidate_name", "N/A")),
        ("Job Title:", metadata.get("job_title", "N/A")),
        ("Interview Duration:", f"{metadata.get('time_limit_minutes', 'N/A')} minutes"),
        ("Generated:", metadata.get("generated_at", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))),
    ]

    for label, value in metadata_items:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        row += 1

    # Interview Overview
    row += 1
    ws[f'A{row}'] = "Interview Objectives"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    objectives = interview_plan.get("interview_overview", {}).get("objectives", [])
    if isinstance(objectives, list):
        for obj in objectives:
            ws[f'A{row}'] = f"• {obj}"
            row += 1
    elif isinstance(objectives, str):
        ws[f'A{row}'] = objectives
        row += 1

    # Time Allocation
    row += 1
    ws[f'A{row}'] = "Time Allocation"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    # Headers for time allocation table
    ws[f'A{row}'] = "Topic"
    ws[f'B{row}'] = "Time (min)"
    ws[f'C{row}'] = "Priority"
    for col in ['A', 'B', 'C']:
        ws[f'{col}{row}'].font = Font(bold=True)
        ws[f'{col}{row}'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    row += 1

    topics = interview_plan.get("prioritized_topics", [])
    for topic in topics:
        ws[f'A{row}'] = topic.get("topic_name", "N/A")
        ws[f'B{row}'] = topic.get("allocated_time_minutes", "N/A")
        ws[f'C{row}'] = topic.get("priority", "N/A")
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 40

def create_questions_sheet(wb, interview_plan):
    """Create questions sheet with all interview questions.

    Returns:
        dict: Mapping of topic names to score cell ranges for formulas
    """
    ws = wb.create_sheet("Questions")

    # Header
    ws.merge_cells('A1:E1')
    ws['A1'] = "INTERVIEW QUESTIONS"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3
    topic_score_ranges = {}  # Store score ranges for each topic

    topics = interview_plan.get("topics_to_cover", [])
    if not topics:
        # Try to get from prioritized_topics
        topics = interview_plan.get("prioritized_topics", [])

    for topic in topics:
        # Topic header
        ws.merge_cells(f'A{row}:E{row}')
        topic_name = topic.get("topic_name", topic.get("topic", "Topic"))
        ws[f'A{row}'] = topic_name.upper()
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal="left", vertical="center")
        row += 1

        # Column headers
        headers = ["#", "Question", "What to Look For", "Follow-up", "Score (1-5)"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(wrap_text=True)
        row += 1

        # Questions
        questions = topic.get("questions", [])
        start_row = row  # Mark start of score range

        for q_idx, question in enumerate(questions, start=1):
            ws.cell(row=row, column=1, value=q_idx)
            ws.cell(row=row, column=2, value=question.get("question", "N/A"))
            ws.cell(row=row, column=3, value=question.get("what_to_look_for", ""))
            ws.cell(row=row, column=4, value=question.get("follow_up", ""))
            ws.cell(row=row, column=5, value="")  # Empty for scoring

            # Set row height and wrap text
            ws.row_dimensions[row].height = 50
            for col in range(1, 6):
                ws.cell(row=row, column=col).alignment = Alignment(wrap_text=True, vertical="top")

            row += 1

        # Store the range of score cells for this topic (column E)
        end_row = row - 1
        if questions:  # Only if there are questions
            topic_score_ranges[topic_name] = f"Questions!E{start_row}:E{end_row}"

        row += 1  # Space between topics

    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12

    return topic_score_ranges

def create_evaluation_sheet(wb, interview_plan, topic_score_ranges):
    """Create evaluation sheet with scoring rubric.

    Args:
        wb: Workbook object
        interview_plan: Interview plan dict
        topic_score_ranges: Dict mapping topic names to score cell ranges from Questions sheet
    """
    ws = wb.create_sheet("Evaluation")

    # Header
    ws.merge_cells('A1:F1')
    ws['A1'] = "CANDIDATE EVALUATION FORM"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3

    # Evaluation rubric
    rubric = interview_plan.get("evaluation_rubric", {})

    # Table headers
    headers = ["Topic", "Weight %", "Score (1-5)", "Notes", "Weighted Score"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row += 1

    topics = interview_plan.get("prioritized_topics", [])
    total_weight = sum([t.get("priority", 1) for t in topics])

    for topic in topics:
        topic_name = topic.get("topic_name", "N/A")
        weight = (topic.get("priority", 1) / total_weight * 100) if total_weight > 0 else 0

        ws.cell(row=row, column=1, value=topic_name)
        # Store weight as number (not text with %) for proper formula calculation
        ws.cell(row=row, column=2, value=weight / 100)  # Store as decimal (0.20 for 20%)
        ws.cell(row=row, column=2).number_format = '0.0%'  # Format as percentage display

        # Score column - auto-calculate average from Questions sheet
        if topic_name in topic_score_ranges:
            # Use IFERROR with AVERAGE to handle empty cells gracefully
            ws.cell(row=row, column=3, value=f"=IFERROR(AVERAGE({topic_score_ranges[topic_name]}),\"\")")
        else:
            ws.cell(row=row, column=3, value="")  # Manual entry if no questions

        ws.cell(row=row, column=4, value="")  # Notes to be filled
        # Weighted score: (Score/5) * Weight% * 100 to get percentage from 0 to Weight%
        # Example: Score=4, Weight=20% -> (4/5) * 0.20 * 100 = 16%
        # Use IF and ISNUMBER to handle empty score cells properly
        ws.cell(row=row, column=5, value=f"=IF(OR(ISBLANK(C{row}),NOT(ISNUMBER(C{row}))),\"\",(C{row}/5)*B{row}*100)")
        ws.cell(row=row, column=5).number_format = '0.0"%"'  # Format as percentage

        row += 1

    # Total row
    ws.cell(row=row, column=1, value="TOTAL SCORE")
    ws.cell(row=row, column=1).font = Font(bold=True)
    # Sum all weighted scores - result will be percentage from 0 to 100%
    # Use IFERROR to handle cases where all scores are empty
    start_eval_row = 4  # First data row in evaluation sheet
    ws.cell(row=row, column=5, value=f"=IFERROR(SUM(E{start_eval_row}:E{row-1}),\"\")")
    ws.cell(row=row, column=5).font = Font(bold=True)
    ws.cell(row=row, column=5).fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    ws.cell(row=row, column=5).number_format = '0.0"%"'  # Format as percentage (0-100%)

    row += 2

    # Decision section
    ws.merge_cells(f'A{row}:B{row}')
    ws[f'A{row}'] = "RECOMMENDATION:"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws.cell(row=row, column=3, value="☐ Hire  ☐ No Hire  ☐ Maybe")
    ws.cell(row=row, column=3).font = Font(size=11)

    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['E'].width = 15

def create_code_challenges_sheet(wb, code_challenges):
    """Create code challenges sheet."""
    ws = wb.create_sheet("Code Challenges")

    # Header
    ws.merge_cells('A1:D1')
    ws['A1'] = "CODE CHALLENGES"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3

    # Coding challenges
    challenges = code_challenges.get("coding_challenges", [])
    for idx, challenge in enumerate(challenges, start=1):
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = f"Challenge {idx}: {challenge.get('metadata', {}).get('difficulty', 'N/A').upper()}"
        ws[f'A{row}'].font = Font(size=12, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        row += 1

        # Problem description
        ws[f'A{row}'] = "Problem:"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'B{row}'] = challenge.get("problem_description", "N/A")
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 60
        row += 1

        # Duration
        ws[f'A{row}'] = "Duration:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = f"{challenge.get('metadata', {}).get('duration_minutes', 'N/A')} minutes"
        row += 1

        # Evaluation criteria
        ws[f'A{row}'] = "Evaluation:"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'B{row}:D{row}')
        criteria = challenge.get("evaluation_criteria", [])
        if isinstance(criteria, list):
            ws[f'B{row}'] = "\n".join([f"• {c}" for c in criteria])
        else:
            ws[f'B{row}'] = str(criteria)
        ws[f'B{row}'].alignment = Alignment(wrap_text=True)
        row += 2

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

def create_notes_sheet(wb, interview_plan):
    """Create notes sheet for interviewer."""
    ws = wb.create_sheet("Notes")

    # Header
    ws.merge_cells('A1:C1')
    ws['A1'] = "INTERVIEWER NOTES"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3

    # Red flags section
    ws[f'A{row}'] = "RED FLAGS TO WATCH FOR:"
    ws[f'A{row}'].font = Font(size=12, bold=True, color="C00000")
    ws[f'A{row}'].fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
    row += 1

    red_flags = interview_plan.get("red_flags_to_watch_for", [])
    for flag in red_flags:
        ws[f'A{row}'] = f"⚠ {flag}"
        ws[f'A{row}'].font = Font(color="C00000")
        row += 1

    row += 2

    # Free-form notes
    ws[f'A{row}'] = "GENERAL NOTES:"
    ws[f'A{row}'].font = Font(size=12, bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    row += 1

    # Add empty rows for notes
    for i in range(20):
        ws.row_dimensions[row + i].height = 20

    # Set column widths
    ws.column_dimensions['A'].width = 80
